import openpyxl

# 读取Excel文件夹 要关闭 xlsx新的excel格式
# windows 下面的路径 有反斜杠 r
# 得到一个workbook对象
wb = openpyxl.load_workbook(r'd:\case_data.xlsx')

# 不直接去获取_sheets属性 私有属性
# print(wb._sheets)
# active 被选择的 激活的
# active_sheet = wb.active

# sheetnames列表当中储存的字符串 和 _sheets 存储对象区别


# 获取所有表单的正确用法
# wb.worksheets

# 获取某一个表单 1、通过索引去获取
sheet = wb.worksheets[0]
# print(sheet)

# 2、通过表单名称获取 sheet1
# sh = wb.get_sheet_by_name('Sheet1') 过时
# print(sh)
# sh = wb['Sheet1']
# print(sh)
# pycharm 不支持sheet. 方法提示

# 读取单个单元格
# 行列从1开始
cell = sheet.cell(1, 2)
print(cell.value)


# 获取某一行 的值
print(sheet[1])
for column in sheet[1]:
    print(column.value)


# 获取某一列
# print(sheet['B'])
#
# 获取多行 1-3行
# print(sheet[1:3])

# 获取所有数据
total_data = list(sheet.rows)
print(total_data)

for row in total_data:
    for cell in row:
        print(cell.value)

# 写入-保存
# cell.value = xxx
# wb.save("文件名称")

# 关闭
wb.close()


