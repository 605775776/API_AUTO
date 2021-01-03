from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

"""
1、打开表单
2、读取标题
3、读取所有的数据（类型是？ 列表嵌入字典）
4、指定单元格写入数据（使用静态方法，不要使用实例方法）

"""


class ExcelHandler():

    """操作excel"""
    def __init__(self, file, sheet_name):

        self.file = file
        self.sheet_name = sheet_name

    def open_sheet(self) -> Worksheet:
        """
        在函数或者方法后面 加->类型 表示此函数返回值是一个这样的类型
        函数注解 这样会有提示
        :return:
        """
        wb = load_workbook(self.file)
        sheet = wb[self.sheet_name]
        return sheet

    # 获取表头
    def header(self):
        sheet = self.open_sheet()
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read(self):
        """
        读取所有的数据
        :return:
        """
        sheet = self.open_sheet()
        rows = list(sheet.rows)
        # rows = list(sheet.rows)[1:]  # 生成器
        # print(type(rows))
        # print(rows)
        """
        列表玩法
        """
        # data = []
        # for row in rows:
        #     row_data = []
        #     # print(row)
        #     for cell in row:
        #         row_data.append(cell.value)
        #     data.append(row_data)
        # return data
        """
        字典玩法
        """
        data = []
        for row in rows[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)

                # 列表转字典 要和header去zip
                data_dict = dict(zip(self.header(), row_data))
            data.append(data_dict)
        return data

    # 实例方法每次都要重新打开文件，用静态方法传参少
    def write_data(self, row, column, data):
        """写入数据"""
        sheet = self.open_sheet()
    @staticmethod
    def write(file, sheet_name, row, column, data):
        wb = load_workbook(file)
        sheet = wb[sheet_name]

        sheet.cell(row, column).value = data
        wb.save(file)
        wb.close()


# if __name__ == '__main__':
#     excel = ExcelHandler(r'd:\case_data.xlsx', 'Sheet1')
#     excel.write(r'd:\case_data.xlsx', 'Sheet1', 3, 1, 'new_value')
    # print(data)
    # print(data)









